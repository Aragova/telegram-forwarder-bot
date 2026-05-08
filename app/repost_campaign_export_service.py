from __future__ import annotations

import csv
import io
import json


def _clean_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, int):
        return str(value)
    return str(value).replace("\r\n", "\n").replace("\r", "\n")


def _csv_bytes(headers: list[str], rows: list[list[object]]) -> bytes:
    buf = io.StringIO(newline="")
    writer = csv.writer(buf, delimiter=";", lineterminator="\n")
    writer.writerow(headers)
    for row in rows:
        writer.writerow([_clean_text(x) for x in row])
    return buf.getvalue().encode("utf-8-sig")


def _format_views(item: dict) -> int | None:
    value = item.get("views")
    if value is None:
        value = item.get("views_total")
    if value is None:
        return None
    return int(value)


def _format_message_ids(item: dict) -> str:
    candidates = [item.get("sent_message_ids"), item.get("sent_message_ids_json")]
    for value in candidates:
        if isinstance(value, list):
            return ",".join(str(int(x)) for x in value if x is not None)
        if isinstance(value, str):
            try:
                parsed = json.loads(value)
            except Exception:
                return _clean_text(value)
            if isinstance(parsed, list):
                return ",".join(str(int(x)) for x in parsed if x is not None)
            return _clean_text(value)
    sent_message_id = item.get("sent_message_id")
    if sent_message_id is not None:
        return str(sent_message_id)
    return ""

def _human_status(value: object) -> str:
    raw = str(value or "").strip().lower()
    mapping = {
        "ok": "есть данные",
        "available": "есть данные",
        "unavailable": "нет данных",
        "problem": "проблема",
        "failed": "ошибка",
        "pending": "ожидает",
        "processing": "обрабатывается",
        "sent": "отправлено",
        "sending": "отправляется",
        "deleted": "удалено",
        "none": "",
        "final_snapshot": "финальный снимок",
        "live": "live-данные",
        "telethon_live": "live-данные",
    }
    return mapping.get(raw, _clean_text(value))


def _human_views_source(value: object) -> str:
    return _human_status(value)


def build_campaign_run_report_csv(report: dict) -> bytes:
    headers = [
        "ID запуска", "ID поста", "Канал/группа", "Telegram ID", "ID темы", "Статус отправки", "Статус удаления",
        "ID сообщения", "ID сообщений", "Просмотры", "Статус просмотров", "Источник данных", "Отправлено",
        "Удалить после", "Удалено", "Проблема", "Ссылка",
    ]
    rows: list[list[object]] = []
    for item in (report or {}).get("items") or []:
        sent_ids_text = _format_message_ids(item)
        rows.append([
            report.get("run_id"),
            report.get("saved_post_id"),
            item.get("target_title"),
            item.get("target_id"),
            item.get("target_thread_id"),
            _human_status(item.get("send_status")),
            _human_status(item.get("delete_status")),
            item.get("sent_message_id"),
            sent_ids_text,
            _format_views(item),
            _human_status(item.get("views_status") or item.get("views_final_status")),
            _human_views_source(item.get("views_source") or ("final_snapshot" if item.get("final_snapshot") else None)),
            item.get("sent_at"),
            item.get("delete_after_at"),
            item.get("deleted_at"),
            item.get("error_text") or item.get("send_error_text") or item.get("delete_error_text"),
            item.get("message_url"),
        ])
    return _csv_bytes(headers, rows)


def build_campaign_run_report_txt(report: dict) -> str:
    items = (report or {}).get("items") or []
    lines = [
        "📊 Отчёт рекламной кампании",
        "",
        f"Запуск: #{int(report.get('run_id') or 0)}",
        f"Пост: #{int(report.get('saved_post_id') or 0)}",
        f"Всего просмотров: {int(report.get('views_total') or 0):,}".replace(",", " "),
        f"Каналов с данными: {int(report.get('views_available') or 0)} / {len(items)}",
        "",
        "Каналы:",
    ]
    for index, item in enumerate(items, 1):
        title = str(item.get("target_title") or item.get("target_id") or "Канал")
        views = _format_views(item)
        if views is None or str(item.get("views_status") or "").lower() != "ok":
            reason = item.get("error_text") or item.get("send_error_text") or item.get("delete_error_text") or item.get("views_status") or "нет данных"
            lines.append(f"{index}. {title} — нет данных — {reason}")
            continue
        status = str(item.get("delete_status") or "").lower()
        suffix = ""
        if status == "deleted":
            suffix = " — удалено"
        message_ids = _format_message_ids(item)
        message_ids_suffix = f" — сообщения: {message_ids}" if message_ids else ""
        lines.append(f"{index}. {title} — {views:,} просмотров{suffix}{message_ids_suffix}".replace(",", " "))
    return "\n".join(lines)


def build_campaign_post_stats_csv(stats: dict) -> bytes:
    headers = ["ID поста", "Канал/группа", "Telegram ID", "Просмотры", "Статус просмотров", "Источник данных", "Запусков", "Без данных", "Проблема"]
    rows = []
    for item in (stats or {}).get("channels_stats") or []:
        rows.append([
            stats.get("saved_post_id"),
            item.get("target_title"),
            item.get("target_id"),
            item.get("views_total"),
            _human_status(item.get("views_status")),
            _human_views_source(item.get("views_source")),
            item.get("runs_count"),
            item.get("unavailable_count"),
            item.get("error_text"),
        ])
    return _csv_bytes(headers, rows)


def build_campaign_post_stats_txt(stats: dict) -> str:
    items = (stats or {}).get("channels_stats") or []
    lines = [
        "📄 Статистика рекламного поста",
        "",
        f"Пост: #{int(stats.get('saved_post_id') or 0)}",
        f"Всего просмотров: {int(stats.get('views_total') or 0):,}".replace(",", " "),
        f"Запусков: {int(stats.get('runs_count') or 0)}",
        f"Размещений: {int(stats.get('placements_sent') or 0)}",
        "",
        "Каналы:",
    ]
    for index, item in enumerate(items, 1):
        title = str(item.get("target_title") or item.get("target_id") or "Канал")
        status = str(item.get("views_status") or "").lower()
        if status == "ok":
            lines.append(f"{index}. {title} — {int(item.get('views_total') or 0):,} просмотров".replace(",", " "))
        else:
            lines.append(f"{index}. {title} — нет данных")
    return "\n".join(lines)

def _style_worksheet(ws, *, wrap_columns: set[str], number_columns: set[str]) -> None:
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
    max_row = ws.max_row
    max_col = ws.max_column
    if max_row <= 0 or max_col <= 0:
        return
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    headers = [cell.value for cell in ws[1]]
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for col_idx, header in enumerate(headers, 1):
        header_text = str(header or "")
        width = len(header_text)
        for row_idx in range(2, max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if header_text in number_columns and cell.value not in (None, ""):
                cell.value = int(cell.value)
            if header_text in wrap_columns:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            value_len = len(str(cell.value or ""))
            if value_len > width:
                width = value_len
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = min(45, max(10, width + 2))


def build_campaign_run_report_xlsx(report: dict) -> bytes:
    from openpyxl import Workbook
    headers = [
        "ID запуска", "ID поста", "Канал/группа", "Telegram ID", "ID темы", "Статус отправки", "Статус удаления",
        "ID сообщения", "ID сообщений", "Просмотры", "Статус просмотров", "Источник данных", "Отправлено",
        "Удалить после", "Удалено", "Проблема", "Ссылка",
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт запуска"
    ws.append(headers)
    for item in (report or {}).get("items") or []:
        ws.append([
            report.get("run_id"),
            report.get("saved_post_id"),
            _clean_text(item.get("target_title")),
            _clean_text(item.get("target_id")),
            _clean_text(item.get("target_thread_id")),
            _human_status(item.get("send_status")),
            _human_status(item.get("delete_status")),
            _clean_text(item.get("sent_message_id")),
            _format_message_ids(item),
            _format_views(item),
            _human_status(item.get("views_status") or item.get("views_final_status")),
            _human_views_source(item.get("views_source") or ("final_snapshot" if item.get("final_snapshot") else None)),
            _clean_text(item.get("sent_at")),
            _clean_text(item.get("delete_after_at")),
            _clean_text(item.get("deleted_at")),
            _clean_text(item.get("error_text") or item.get("send_error_text") or item.get("delete_error_text")),
            _clean_text(item.get("message_url")),
        ])
    _style_worksheet(
        ws,
        wrap_columns={"Проблема", "Ссылка"},
        number_columns={"ID запуска", "ID поста", "ID темы", "ID сообщения", "Просмотры"},
    )
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def build_campaign_post_stats_xlsx(stats: dict) -> bytes:
    from openpyxl import Workbook
    headers = ["ID поста", "Канал/группа", "Telegram ID", "Просмотры", "Статус просмотров", "Источник данных", "Запусков", "Без данных", "Проблема"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Статистика поста"
    ws.append(headers)
    for item in (stats or {}).get("channels_stats") or []:
        ws.append([
            stats.get("saved_post_id"),
            _clean_text(item.get("target_title")),
            _clean_text(item.get("target_id")),
            item.get("views_total"),
            _human_status(item.get("views_status")),
            _human_views_source(item.get("views_source")),
            item.get("runs_count"),
            item.get("unavailable_count"),
            _clean_text(item.get("error_text")),
        ])
    _style_worksheet(
        ws,
        wrap_columns={"Проблема"},
        number_columns={"ID поста", "Просмотры", "Запусков", "Без данных"},
    )
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
